from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
import uuid
from pathlib import Path
from datetime import datetime
import os
from app.models import DeliveryDate, DeliveryPointModel

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from collections import defaultdict

from app.config import ADMIN_USERNAME, ADMIN_PASSWORD, PUBLIC_BACKEND_URL
from app.db import get_db
from app.schemas import (
    AdminLoginRequest,
    AdminLoginResponse,
    ShopStatusResponse,
    ShopStatusUpdateRequest,
    ShopStatusUpdateResponse,
    AdminCartsListResponse,
    AdminCartResponse,
    AdminProductTotalsResponse,
    AdminProductsListResponse,
    AdminProductResponse,
    AdminProductCreateRequest,
    AdminDeliveryPointsListResponse,
    AdminDeliveryPointResponse,
    AdminDeliveryPointCreateRequest,
    AdminProductUpdateRequest,
    DeleteResponse,
    AdminDeliveryPointUpdateRequest,
    UploadImageResponse,
    ClearCartsResponse,
    DeliveryDateResponse,
    DeliveryDateCreateRequest,
    DeliveryDateUpdateRequest,
)
from app.services.db_service import DBService
from app.utils.admin_auth import create_admin_access_token, verify_admin_token

CLIENT_EXPORT_COLUMNS = [
("FRESH-FISH-CARP01","карп,шт"),
("FRESH-FISH-STURGEON01","Осётр /Stör,,штуки"),
("FILLET01","Фарш лакса норв."),
("FROZEN01","котлети"),
("FROZEN04","морской коктейль"),
("SHRIMP02",'Креветки "Black Tiger"'),
("FROZEN03","мидии"),
("SEAWEED01","вакаме"),
("FROZEN02","крабовые палочки, кг"),
("MATJES01","Matjes file ... КРУГЛОЕ ВЕДЕРКО"),
("MATJES02","Matjes file ... КВАДРАТНОЕ ВЕДЕРКО"),
("FRESH-FISH-COD01","тушка без головы"),
("FRESH-FISH-CARP03","караси"),
("FRESH-FISH-PERCH01","морской окунь"),
("FRESH-FISH-SALMON03","лакс нарезной"),
("FRESH-FISH-SALMON04","lachs razdelka"),
("FILLET03","Филе лакса свежее"),
("FILLET04","Филе Seelachs"),
("FROZEN08","кальмары тушки"),
("FROZEN10","камбала"),
("FROZEN11","палтус"),
("FROZEN12","пангасиус"),
("FROZEN13","хек"),
("FROZEN14","Лакс норвежский цельный"),
("FROZEN15","мясо речных раков"),
("FROZEN16","щупальца осьминога"),
("SHRIMP04","Креветки аргентинские"),
("SMOKED01","Угорь гор копч"),
("SMOKED02","Лакс хол копч"),
("SMOKED03","Скумбрия хол копч"),
("SMOKED04","Дораде гор копч"),
("SMOKED05","Брюшки лакса"),
("SMOKED06","нерка"),
("CAVIAR03","икра 100"),
("CAVIAR04","икра форели"),
("CAVIAR05","икра 500"),
("CAVIAR06","икра кеты"),
("SHRIMP05","Креветки отварные 5 кг"),
]

router = APIRouter(prefix="/admin", tags=["admin"])

def _sheet_title_from_delivery_date(value: str | None) -> str:
    if not value:
        return "Без даты"

    text = str(value).strip()
    if not text:
        return "Без даты"

    # Excel sheet title max 31 chars and cannot contain some symbols
    safe = text.replace("/", ".").replace("\\", ".").replace(":", "-").replace("?", "")
    safe = safe.replace("*", "").replace("[", "(").replace("]", ")")
    return safe[:31]


def _group_submitted_carts_by_delivery_date(carts: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = {}

    for cart in carts:
        if cart.get("status") not in {"submitted", "locked"}:
            continue

        key = cart.get("delivery_date") or "Без даты"
        grouped.setdefault(key, []).append(cart)

    for key in grouped:
        grouped[key].sort(
            key=lambda cart: (
                cart.get("city") or "",
                cart.get("delivery_point") or "",
                cart.get("customer_name") or "",
            )
        )

    return grouped

@router.post("/login", response_model=AdminLoginResponse)
def admin_login(payload: AdminLoginRequest):

    if payload.username != ADMIN_USERNAME or payload.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    access_token = create_admin_access_token()

    return AdminLoginResponse(
        access_token=access_token,
        token_type="bearer"
    )


@router.get("/shop-status", response_model=ShopStatusResponse, dependencies=[Depends(verify_admin_token)])
def get_admin_shop_status(db: Session = Depends(get_db)):
    service = DBService(db)
    return ShopStatusResponse(status=service.get_shop_status())

@router.get("/carts", response_model=AdminCartsListResponse, dependencies=[Depends(verify_admin_token)])
def get_admin_carts(
    delivery_date: str | None = None,
    db: Session = Depends(get_db),
):
    service = DBService(db)
    carts = service.get_all_carts(delivery_date=delivery_date)
    return AdminCartsListResponse(carts=carts)


@router.get("/carts/{cart_id}", response_model=AdminCartResponse, dependencies=[Depends(verify_admin_token)])
def get_admin_cart(cart_id: int, db: Session = Depends(get_db)):
    service = DBService(db)
    return service.get_cart_by_id(cart_id)

@router.get(
    "/product-totals",
    response_model=AdminProductTotalsResponse,
    dependencies=[Depends(verify_admin_token)],
)
def get_admin_product_totals(db: Session = Depends(get_db)):
    service = DBService(db)
    totals = service.get_product_totals()
    return AdminProductTotalsResponse(totals=totals)


@router.get(
    "/export/orders.xlsx",
    dependencies=[Depends(verify_admin_token)],
)
def export_orders_excel(db: Session = Depends(get_db)):
    service = DBService(db)
    carts = service.get_all_carts()
    totals = service.get_product_totals()

    wb = Workbook()

    grouped = _group_submitted_carts_by_delivery_date(carts)

    if not grouped:
        ws = default_ws
        ws.append(["Нет заказов"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=fishshop_client_format.xlsx"
            },
        )

    # --- Общий лист со всеми заказами ---
    ws_orders = wb.active
    ws_orders.title = "Orders"
    ws_orders.append([
        "№ заказа",
        "Имя",
        "Телефон",
        "Telegram",
        "Город",
        "Точка выдачи",
        "Дата выдачи",
        "Время",
        "Состав заказа",
        "Комментарий",
        "Обновлено",
    ])

    for delivery_date in sorted(grouped.keys()):
        for cart in grouped[delivery_date]:
            items_text = "; ".join(
                [
                    f"{item['product_name']} ({item['sku']}) x {item['qty']} {item['unit']}".strip()
                    for item in cart["items"]
                ]
            )

            ws_orders.append([
                cart.get("pickup_number", ""),
                cart.get("customer_name", ""),
                cart.get("phone", ""),
                cart.get("telegram_username") or cart.get("telegram_user_id", ""),
                cart.get("city", ""),
                cart.get("delivery_point", ""),
                cart.get("delivery_date", ""),
                cart.get("approx_time", "") or "",
                items_text,
                cart.get("comment", ""),
                cart.get("updated_at", ""),
            ])

    # --- Отдельные листы по датам ---
    for delivery_date in sorted(grouped.keys()):
        sheet_name = _sheet_title_from_delivery_date(delivery_date)
        ws_date = wb.create_sheet(title=sheet_name)

        ws_date.append([
            "№ заказа",
            "Имя",
            "Телефон",
            "Telegram",
            "Город",
            "Точка выдачи",
            "Дата выдачи",
            "Время",
            "Состав заказа",
            "Комментарий",
            "Обновлено",
        ])

        for cart in grouped[delivery_date]:
            items_text = "; ".join(
                [
                    f"{item['product_name']} ({item['sku']}) x {item['qty']} {item['unit']}".strip()
                    for item in cart["items"]
                ]
            )

            ws_date.append([
                cart.get("pickup_number", ""),
                cart.get("customer_name", ""),
                cart.get("phone", ""),
                cart.get("telegram_username") or cart.get("telegram_user_id", ""),
                cart.get("city", ""),
                cart.get("delivery_point", ""),
                cart.get("delivery_date", ""),
                cart.get("approx_time", "") or "",
                items_text,
                cart.get("comment", ""),
                cart.get("updated_at", ""),
            ])

    # --- Totals ---
    ws_totals = wb.create_sheet(title="Totals")
    ws_totals.append([
        "SKU",
        "Товар",
        "Ед.",
        "Общее количество",
    ])

    for item in totals:
        ws_totals.append([
            item["sku"],
            item["product_name"],
            item["unit"],
            item["total_qty"],
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=fishshop_orders.xlsx"
        },
    )

@router.get(
    "/products",
    response_model=AdminProductsListResponse,
    dependencies=[Depends(verify_admin_token)],
)
def get_admin_products(db: Session = Depends(get_db)):
    service = DBService(db)
    products = service.get_all_products()
    return AdminProductsListResponse(products=products)

@router.put(
    "/products/{product_id}",
    response_model=AdminProductResponse,
    dependencies=[Depends(verify_admin_token)],
)
def update_admin_product(
    product_id: int,
    payload: AdminProductUpdateRequest,
    db: Session = Depends(get_db),
):
    service = DBService(db)
    return service.update_product(product_id, payload)

@router.post(
    "/products",
    response_model=AdminProductResponse,
    dependencies=[Depends(verify_admin_token)],
)
def create_admin_product(payload: AdminProductCreateRequest, db: Session = Depends(get_db)):
    service = DBService(db)
    return service.create_product(payload)


@router.post(
    "/products/{product_id}/toggle",
    response_model=AdminProductResponse,
    dependencies=[Depends(verify_admin_token)],
)
def toggle_admin_product(product_id: int, db: Session = Depends(get_db)):
    service = DBService(db)
    return service.toggle_product_active(product_id)


@router.get(
    "/delivery-points",
    response_model=AdminDeliveryPointsListResponse,
    dependencies=[Depends(verify_admin_token)],
)
def get_admin_delivery_points(db: Session = Depends(get_db)):
    service = DBService(db)
    points = service.get_all_delivery_points()
    return AdminDeliveryPointsListResponse(delivery_points=points)


@router.post(
    "/delivery-points",
    response_model=AdminDeliveryPointResponse,
    dependencies=[Depends(verify_admin_token)],
)
def create_admin_delivery_point(payload: AdminDeliveryPointCreateRequest, db: Session = Depends(get_db)):
    service = DBService(db)
    return service.create_delivery_point(payload)


@router.post(
    "/delivery-points/{point_id}/toggle",
    response_model=AdminDeliveryPointResponse,
    dependencies=[Depends(verify_admin_token)],
)
def toggle_admin_delivery_point(point_id: int, db: Session = Depends(get_db)):
    service = DBService(db)
    return service.toggle_delivery_point_active(point_id)

@router.delete(
    "/products/{product_id}",
    response_model=DeleteResponse,
    dependencies=[Depends(verify_admin_token)],
)
def delete_admin_product(product_id: int, db: Session = Depends(get_db)):
    service = DBService(db)
    return service.delete_product(product_id)

@router.put(
    "/delivery-points/{point_id}",
    response_model=AdminDeliveryPointResponse,
    dependencies=[Depends(verify_admin_token)],
)
def update_admin_delivery_point(
    point_id: int,
    payload: AdminDeliveryPointUpdateRequest,
    db: Session = Depends(get_db),
):
    service = DBService(db)
    return service.update_delivery_point(point_id, payload)


@router.delete(
    "/delivery-points/{point_id}",
    response_model=DeleteResponse,
    dependencies=[Depends(verify_admin_token)],
)
def delete_admin_delivery_point(point_id: int, db: Session = Depends(get_db)):
    service = DBService(db)
    return service.delete_delivery_point(point_id)

@router.post(
    "/upload/product-image",
    response_model=UploadImageResponse,
    dependencies=[Depends(verify_admin_token)],
)
async def upload_product_image(file: UploadFile = File(...)):
    allowed_types = {"image/jpeg", "image/png", "image/webp", "image/jpg"}
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Unsupported file type")

    ext = Path(file.filename or "").suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".webp"}:
        ext = ".jpg"

    filename = f"{uuid.uuid4().hex}{ext}"
    upload_dir = Path("uploads/products")
    upload_dir.mkdir(parents=True, exist_ok=True)

    file_path = upload_dir / filename

    content = await file.read()
    file_path.write_bytes(content)

    image_url = f"{PUBLIC_BACKEND_URL}/uploads/products/{filename}"
    return UploadImageResponse(image_url=image_url)

@router.post(
    "/clear-carts",
    response_model=ClearCartsResponse,
    dependencies=[Depends(verify_admin_token)],
)
def clear_admin_carts(db: Session = Depends(get_db)):
    service = DBService(db)
    result = service.clear_all_carts()
    return ClearCartsResponse(**result)



@router.get(
    "/export/client-format.xlsx",
    dependencies=[Depends(verify_admin_token)],
)
def export_client_format_excel(db: Session = Depends(get_db)):
    service = DBService(db)
    carts = service.get_all_carts()
    mapping = service.get_client_export_mapping()

    grouped = _group_submitted_carts_by_delivery_date(carts)

    wb = Workbook()
    # Удаляем пустой стартовый лист, чтобы не мешал
    default_ws = wb.active
    default_ws.title = "Client Format"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    base_headers = ["Nr.", "Kunde", "Auto 1", "Zeit", "Goroda", "Gebiet"]

    for delivery_date in sorted(grouped.keys()):
        ws = wb.create_sheet(title=_sheet_title_from_delivery_date(delivery_date))

        headers = base_headers + [entry["column_name"] for entry in mapping]
        ws.append(headers)

        weight_row = ["", "", "", "", "", "Gewicht"] + [entry.get("weight_label", "") for entry in mapping]
        ws.append(weight_row)

        price_row = ["", "", "", "", "", "Preis"] + [entry.get("price_label", "") for entry in mapping]
        ws.append(price_row)

        row_index = 1

        for cart in grouped[delivery_date]:
            items_map = {}
            for item in cart["items"]:
                sku = (item.get("sku") or "").strip()
                qty = float(item.get("qty") or 0)
                items_map[sku] = qty

            row = [
                cart.get("pickup_number", ""),
                cart.get("customer_name", ""),
                "",
                cart.get("approx_time", "") or row_index,
                cart.get("city", ""),
                cart.get("delivery_point", ""),
            ]

            for entry in mapping:
                qty = items_map.get(entry["sku"], "")
                row.append("" if qty in (0, 0.0) else qty)

            ws.append(row)
            row_index += 1

        header_fills = {
            "Kunde": "FFF200",
            "Auto 1": "D9D9D9",
            "Zeit": "F2F2F2",
            "Goroda": "FFF200",
            "Gebiet": "00FF00",
        }

        for col_idx, cell in enumerate(ws[1], start=1):
            if col_idx <= 6:
                fill_color = header_fills.get(cell.value, "D9D9D9")
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    text_rotation=90,
                    wrap_text=True,
                )
            else:
                entry = mapping[col_idx - 7]
                fill_color = entry.get("fill_color") or "D9D9D9"
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    text_rotation=90,
                    wrap_text=True,
                )

            cell.border = border

            cell.border = border

        for row_num in [2, 3]:
            for cell in ws[row_num]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = Font(bold=False)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        fixed_widths = {
            1: 8,
            2: 18,
            3: 10,
            4: 10,
            5: 16,
            6: 24,
        }

        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            if col_idx in fixed_widths:
                ws.column_dimensions[letter].width = fixed_widths[col_idx]
            else:
                ws.column_dimensions[letter].width = 9

        ws.row_dimensions[1].height = 120
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 22

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=fishshop_client_format.xlsx"
        },
    )

@router.get(
    "/export/client-format/validate",
    dependencies=[Depends(verify_admin_token)],
)
def validate_client_format_export(db: Session = Depends(get_db)):
    service = DBService(db)
    return service.validate_client_export_mapping()

@router.post(
    "/shop-status",
    response_model=ShopStatusUpdateResponse,
    dependencies=[Depends(verify_admin_token)],
)
def update_shop_status(
    payload: ShopStatusUpdateRequest,
    db: Session = Depends(get_db),
):
    service = DBService(db)

    if payload.status == "locked":
        service.finalize_month_orders()

    service.set_shop_status(payload.status)

    return ShopStatusUpdateResponse(
        status=payload.status,
        updated=True,
    )

@router.get(
    "/delivery-points/{point_id}/dates",
    response_model=list[DeliveryDateResponse],
    dependencies=[Depends(verify_admin_token)],
)
def get_point_dates(point_id: int, db: Session = Depends(get_db)):
    point = db.query(DeliveryPointModel).filter(DeliveryPointModel.id == point_id).first()
    if not point:
        raise HTTPException(status_code=404, detail="Точка выдачи не найдена")

    return (
        db.query(DeliveryDate)
        .filter(DeliveryDate.delivery_point_id == point_id)
        .order_by(DeliveryDate.delivery_date.asc())
        .all()
    )

@router.post(
    "/delivery-points/{point_id}/dates",
    response_model=DeliveryDateResponse,
    dependencies=[Depends(verify_admin_token)],
)
def create_point_date(
    point_id: int,
    payload: DeliveryDateCreateRequest,
    db: Session = Depends(get_db),
):
    point = db.query(DeliveryPointModel).filter(DeliveryPointModel.id == point_id).first()
    if not point:
        raise HTTPException(status_code=404, detail="Точка выдачи не найдена")

    conflict = (
        db.query(DeliveryDate)
        .filter(
            DeliveryDate.delivery_point_id == point_id,
            DeliveryDate.delivery_date == payload.delivery_date,
        )
        .first()
    )
    if conflict:
        raise HTTPException(status_code=409, detail="Такая дата для этой точки уже существует")

    item = DeliveryDate(
        delivery_point_id=point_id,
        city=point.city,
        delivery_date=payload.delivery_date,
        approx_time=(payload.approx_time or "").strip() or None,
        active=payload.active,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return item

@router.put(
    "/delivery-dates/{date_id}",
    response_model=DeliveryDateResponse,
    dependencies=[Depends(verify_admin_token)],
)
def update_point_date(
    date_id: int,
    payload: DeliveryDateUpdateRequest,
    db: Session = Depends(get_db),
):
    item = db.query(DeliveryDate).filter(DeliveryDate.id == date_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Дата не найдена")

    conflict = (
        db.query(DeliveryDate)
        .filter(
            DeliveryDate.id != date_id,
            DeliveryDate.delivery_point_id == item.delivery_point_id,
            DeliveryDate.delivery_date == payload.delivery_date,
        )
        .first()
    )
    if conflict:
        raise HTTPException(status_code=409, detail="Такая дата для этой точки уже существует")

    item.delivery_date = payload.delivery_date
    item.approx_time = (payload.approx_time or "").strip() or None
    item.active = payload.active
    item.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(item)
    return item

@router.delete(
    "/delivery-dates/{date_id}",
    dependencies=[Depends(verify_admin_token)],
)
def delete_point_date(date_id: int, db: Session = Depends(get_db)):
    item = db.query(DeliveryDate).filter(DeliveryDate.id == date_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Дата не найдена")

    db.delete(item)
    db.commit()
    return {"deleted": True}